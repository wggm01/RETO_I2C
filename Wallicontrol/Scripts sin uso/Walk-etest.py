import time
import csv
import openpyxl
from openpyxl import load_workbook
import datetime
import machinarie

wb= load_workbook('data4test/vuelta1.xlsx')
sheet= wb['Hoja1']
i=1

#Ciclo repetido
print("Wall-i actualmente se encuentra esperando su instruccion")
instruccion = raw_input("Presion ""y"" y luego enter para empezar\n ")
while instruccion == 'y':

        latitud_raw= sheet.cell(row=i, column=1).value
        longitud_raw= sheet.cell(row=i, column=2).value
	
        i +=1

        nrp = machinarie.not_repeatcoord(latitud_raw,longitud_raw)
        if nrp != None:
            latitud,longitud = nrp
        virtual_0 = machinarie.virtual_pos0(latitud,longitud)
        
       
        latv0,lonv0= virtual_0

        virtual_1 = machinarie.virtual_pos1(latitud,longitud)
        
       
        latv1,lonv1= virtual_1

        virtual_2 = machinarie.virtual_pos2(latitud,longitud)
        
        
        latv2,lonv2= virtual_2

        virtual_3 = machinarie.virtual_pos3(latitud,longitud)
        
        
        latv3,lonv3= virtual_3

        reg0 = machinarie.check_0drp(latitud,longitud,latv0,lonv0)
        reg1 = machinarie.check_1drp(latitud,longitud,latv1,lonv1)
        reg2 = machinarie.check_2drp(latitud,longitud,latv2,lonv2)
        reg3 = machinarie.check_3drp(latitud,longitud,latv3,lonv3)
        

	time.sleep(.5)